import json
import os
import time
from groq import Groq
from dotenv import load_dotenv

load_dotenv()

_client = None

def configure_groq(api_key: str = None):
    """Configure the Groq API client."""
    global _client
    key = api_key or os.getenv("GROQ_API_KEY")
    if not key:
        raise ValueError("Please provide a valid Groq API key.")
    _client = Groq(api_key=key)

def analyze_resume(resume_text: str, job_description: str) -> dict:
    """
    Analyze a single resume against a job description using Groq (Llama 3.3 70B).
    
    Args:
        resume_text: Extracted text from the resume PDF
        job_description: The job description text
    
    Returns:
        Dictionary with: candidate_name, match_score, strengths, gaps, recommendation
    """
    global _client
    if not _client:
        configure_groq()

    prompt = f"""You are an expert HR recruiter and resume screening specialist.

Analyze the following resume against the provided job description. Evaluate the candidate thoroughly and provide a structured assessment.

---
JOB DESCRIPTION:
{job_description}
---

---
RESUME:
{resume_text}
---

You MUST respond with ONLY a valid JSON object (no markdown, no code blocks, no extra text) in this exact format:
{{
    "candidate_name": "Full name of the candidate (extract from resume, use 'Unknown' if not found)",
    "match_score": <integer from 0 to 100>,
    "strengths": ["strength 1", "strength 2", "strength 3"],
    "gaps": ["gap 1", "gap 2", "gap 3"],
    "recommendation": "<one of: Strong Fit, Moderate Fit, Not Fit>"
}}

SCORING GUIDELINES:
- 80-100: Strong Fit — candidate meets most/all key requirements
- 50-79: Moderate Fit — candidate meets some requirements but has notable gaps
- 0-49: Not Fit — candidate lacks critical requirements

For strengths and gaps:
- Provide exactly 2-3 points each
- Be specific and reference actual skills/experience from the resume
- Compare directly against job description requirements
- CRITICAL: Do NOT use any emojis in your response. Keep the tone highly professional, clean, and premium.

Respond with ONLY the JSON object, nothing else."""

    max_retries = 3
    for attempt in range(max_retries):
        try:
            response = _client.chat.completions.create(
                model="llama-3.3-70b-versatile",
                messages=[
                    {"role": "system", "content": "You are an expert HR recruiter. Always respond with valid JSON only."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.3,
                max_tokens=1024,
            )
            
            response_text = response.choices[0].message.content.strip()
            
            # Remove markdown code block if present
            if response_text.startswith("```"):
                lines = response_text.split("\n")
                lines = [l for l in lines if not l.strip().startswith("```")]
                response_text = "\n".join(lines)
            
            result = json.loads(response_text)
            
            # Validate and sanitize
            result["match_score"] = max(0, min(100, int(result.get("match_score", 0))))
            result["strengths"] = result.get("strengths", ["N/A"])[:3]
            result["gaps"] = result.get("gaps", ["N/A"])[:3]
            
            if result.get("recommendation") not in ("Strong Fit", "Moderate Fit", "Not Fit"):
                score = result["match_score"]
                if score >= 80:
                    result["recommendation"] = "Strong Fit"
                elif score >= 50:
                    result["recommendation"] = "Moderate Fit"
                else:
                    result["recommendation"] = "Not Fit"
            
            return result

        except json.JSONDecodeError:
            return {
                "candidate_name": "Parse Error",
                "match_score": 0,
                "strengths": ["Could not parse response"],
                "gaps": ["LLM response was not valid JSON"],
                "recommendation": "Not Fit"
            }
        except Exception as e:
            error_msg = str(e)
            if ("429" in error_msg or "rate" in error_msg.lower()) and attempt < max_retries - 1:
                wait_time = 5 * (2 ** attempt)
                time.sleep(wait_time)
                continue
            return {
                "candidate_name": "Error",
                "match_score": 0,
                "strengths": ["Error occurred during analysis"],
                "gaps": [error_msg[:100]],
                "recommendation": "Not Fit"
            }

def rank_candidates(results: list[dict]) -> list[dict]:
    """Sort candidates by match score (highest first) and add rank numbers."""
    sorted_results = sorted(results, key=lambda x: x.get("match_score", 0), reverse=True)
    for i, candidate in enumerate(sorted_results, 1):
        candidate["rank"] = i
    return sorted_results


def generate_excel(results: list[dict], job_description: str = "") -> bytes:
    """
    Generate an Excel file from screening results.

    Args:
        results: List of candidate result dictionaries (should be ranked)
        job_description: Optional job description to include in metadata sheet

    Returns:
        bytes: Excel file as bytes (for download)
    """
    import pandas as pd
    from io import BytesIO

    # Prepare data for main sheet
    excel_data = []
    for r in results:
        excel_data.append({
            "Rank": r.get("rank", ""),
            "Candidate Name": r.get("candidate_name", "Unknown"),
            "File Name": r.get("file_name", ""),
            "Match Score": r.get("match_score", 0),
            "Recommendation": r.get("recommendation", ""),
            "Strength 1": r.get("strengths", [""])[0] if len(r.get("strengths", [])) > 0 else "",
            "Strength 2": r.get("strengths", ["", ""])[1] if len(r.get("strengths", [])) > 1 else "",
            "Strength 3": r.get("strengths", ["", "", ""])[2] if len(r.get("strengths", [])) > 2 else "",
            "Gap 1": r.get("gaps", [""])[0] if len(r.get("gaps", [])) > 0 else "",
            "Gap 2": r.get("gaps", ["", ""])[1] if len(r.get("gaps", [])) > 1 else "",
            "Gap 3": r.get("gaps", ["", "", ""])[2] if len(r.get("gaps", [])) > 2 else ""
        })

    df = pd.DataFrame(excel_data)

    # Create Excel in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Main results sheet
        df.to_excel(writer, sheet_name='Screening Results', index=False)
        
        workbook = writer.book
        worksheet = writer.sheets['Screening Results']
        
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        header_font = Font(name='Segoe UI', bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(left=Side(style='thin', color='D9D9D9'),
                             right=Side(style='thin', color='D9D9D9'),
                             top=Side(style='thin', color='D9D9D9'),
                             bottom=Side(style='thin', color='D9D9D9'))
        cell_alignment = Alignment(vertical='top', wrap_text=True)
        
        # Style headers
        for col_num, cell in enumerate(worksheet[1], 1):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            
        # Style data rows
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = cell_alignment
                cell.border = thin_border
                # Add conditional formatting for "Match Score" column (column D)
                if cell.column == 4 and isinstance(cell.value, (int, float)):
                    if cell.value >= 80:
                        cell.font = Font(color='006100', bold=True)
                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    elif cell.value >= 50:
                        cell.font = Font(color='9C6500', bold=True)
                        cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                    else:
                        cell.font = Font(color='9C0006', bold=True)
                        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                # Add conditional formatting for "Recommendation" column (column E)
                if cell.column == 5:
                    if cell.value == "Strong Fit":
                        cell.font = Font(color='006100', bold=True)
                    elif cell.value == "Moderate Fit":
                        cell.font = Font(color='9C6500', bold=True)
                    else:
                        cell.font = Font(color='9C0006', bold=True)

        # Auto-adjust column widths with a minimum and maximum width
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    lines = str(cell.value).split('\n')
                    longest_line = max([len(line) for line in lines])
                    if longest_line > max_length:
                        max_length = longest_line
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            adjusted_width = max(adjusted_width, 15) # Minimum width 
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Add summary sheet if job description provided
        if job_description:
            summary_data = [
                ["Job Description", job_description],
                ["Total Candidates", len(results)],
                ["Average Score", round(sum(r.get("match_score", 0) for r in results) / len(results), 2) if results else 0],
                ["Strong Fits", sum(1 for r in results if r.get("recommendation") == "Strong Fit")],
                ["Moderate Fits", sum(1 for r in results if r.get("recommendation") == "Moderate Fit")],
                ["Not Fit", sum(1 for r in results if r.get("recommendation") == "Not Fit")],
                ["Generated On", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")]
            ]
            summary_df = pd.DataFrame(summary_data, columns=["Metric", "Value"])
            summary_df.to_excel(writer, sheet_name='Summary', index=False)

            ws = writer.sheets['Summary']
            
            # Style Summary Headers
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
                
            # Style Summary Rows
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = cell_alignment
                    cell.border = thin_border

            # Adjust column widths for summary
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 80

    output.seek(0)
    return output.read()
